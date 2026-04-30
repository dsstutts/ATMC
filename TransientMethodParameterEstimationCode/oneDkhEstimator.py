#!/usr/bin/env python3
"""
oneDkhEstimator.py -- modernized 1-D transient convection parameter estimator.

This is a cleaned-up version of the original 2021 script for the model in

    https://doi.org/10.1016/j.applthermaleng.2020.116362

Major fixes/changes in this version
-----------------------------------
1. Fixes the undefined ``plotname`` bug.
2. Uses pathlib consistently for paths.
3. Writes PDF plots by default to FIGS/.
4. Writes tabular results to CSV and, when openpyxl is available, XLSX.
5. Adds a command-line interface for material/test/file selection.
6. Discovers matching CSV files in the data directory by default and reports
   only the number selected, not a list of expected-but-missing files.
7. Replaces the old xlwt dependency with openpyxl fallback logic.
8. Uses residual RMS and a safer covariance calculation based on the Jacobian.
9. Keeps the original three-stage estimation logic:
      a. finite-difference boundary fit,
      b. steady-state fit for initial k guess,
      c. transient h, k, Pss fit.
10. Vectorizes the Fourier-series transient model over the retained modes.
11. Processes independent data files in parallel; automatic worker selection is
   capped by the number of files so 9 files do not spawn 16 workers.
12. Reads data robustly after arbitrary metadata/header preamble lines.
13. Requires at least 8 numeric columns but safely ignores extra acquisition columns.

Examples
--------
Run all matching aluminum temperature files found in DATA_DIR:

    python oneDkhEstimator.py

Run all available temperature files for Al, Cu, and SS:

    python oneDkhEstimator.py --material all --test temp --data-dir data

Run the built-in legacy expected file list instead of directory discovery:

    python oneDkhEstimator.py --material al --test temp --expected-list

Run just one file with fewer Fourier terms for a fast smoke test:

    python oneDkhEstimator.py --material al --files 50al1 --N 25 --no-show

Notes
-----
The CSV input file is expected to contain at least eight columns:

    time, TC1, TC2, TC3, TC4, TC5, TC6, duty_cycle [, extra columns ...]

Files may include arbitrary metadata/header preamble lines before the numeric
data begin. Numeric data rows are detected by finding rows with at least eight
numeric columns. If future acquisition files contain up to 21 columns, this
script uses the first eight columns above and safely ignores the trailing
columns unless/until they are explicitly mapped. The six thermocouple channels
are assumed to be temperature rises relative to the ambient/reference value.
The ambient Tinf is used only for plotting absolute temperatures.
"""

from __future__ import annotations

import argparse
import csv
import math
import os
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
from scipy.optimize import least_squares


PROGRAM_NAME = "oneDkhEstimator"
VERSION = "2.3.1"


@dataclass(frozen=True)
class MaterialCase:
    key: str
    label: str
    x_tc: tuple[float, float, float, float, float, float]  # thermocouple locations, m
    rho: float                                             # density, kg/m^3
    diameter: float                                        # rod diameter, m
    length: float                                          # rod length, m
    c_p: float                                             # specific heat, J/(kg K)
    plot_time: float                                       # plot x-limit, s
    temp_files: tuple[str, ...]
    temp_v_guess: tuple[float, ...]
    temp_ymax: tuple[float, ...]
    temp_tinf: float
    wind_files: tuple[str, ...]
    wind_v_guess: tuple[float, ...]
    wind_ymax: tuple[float, ...]
    wind_tinf: float


MATERIALS: dict[str, MaterialCase] = {
    "cu": MaterialCase(
        key="cu",
        label="Copper",
        x_tc=(0.000, 0.015, 0.030, 0.045, 0.060, 0.075),
        rho=8912.93,
        diameter=0.003175,
        length=0.150,
        c_p=384.93,
        plot_time=300.0,
        temp_files=("50cu1", "75cu1", "100cu1"),
        temp_v_guess=(5.0, 5.0, 5.0),
        temp_ymax=(60.0, 90.0, 120.0),
        temp_tinf=20.5,
        wind_files=("cu1", "cu2", "cu3", "cu4", "cu5", "cu6", "cu7", "cu8", "cu9", "cu10"),
        wind_v_guess=(1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0),
        wind_ymax=(90.0,) * 10,
        wind_tinf=20.6,
    ),
    "al": MaterialCase(
        key="al",
        label="Aluminum",
        x_tc=(0.000, 0.010, 0.020, 0.030, 0.040, 0.050),
        rho=2767.99,
        diameter=0.003175,
        length=0.150,
        c_p=896.0,
        plot_time=300.0,
        temp_files=("50al1", "50al2", "50al3", "75al1", "75al2", "75al3", "100al1", "100al2", "100al3"),
        temp_v_guess=(5.0,) * 9,
        temp_ymax=(60.0, 60.0, 60.0, 90.0, 90.0, 90.0, 120.0, 120.0, 120.0),
        temp_tinf=20.5,
        wind_files=("al1", "al2", "al3", "al4", "al5", "al6", "al7", "al8", "al9", "al10"),
        wind_v_guess=(1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0),
        wind_ymax=(90.0,) * 10,
        wind_tinf=20.6,
    ),
    "ss": MaterialCase(
        key="ss",
        label="Stainless steel",
        x_tc=(0.000, 0.007, 0.014, 0.021, 0.028, 0.035),
        rho=8030.0,
        diameter=0.003175,
        length=0.150,
        c_p=502.0,
        plot_time=800.0,
        temp_files=("50ss1", "50ss2", "50ss3", "75ss1", "75ss2", "75ss3", "100ss1", "100ss2", "100ss3"),
        temp_v_guess=(5.0,) * 9,
        temp_ymax=(60.0, 60.0, 60.0, 90.0, 90.0, 90.0, 120.0, 120.0, 120.0),
        temp_tinf=20.5,
        wind_files=("ss1", "ss2", "ss3", "ss4", "ss5", "ss6", "ss7", "ss8", "ss9", "ss10"),
        wind_v_guess=(1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0),
        wind_ymax=(90.0,) * 10,
        wind_tinf=21.2,
    ),
}


RESULT_COLUMNS = [
    "version",
    "run_timestamp",
    "material",
    "test",
    "filename",
    "n_rows",
    "n_columns_detected",
    "skipped_preamble_rows",
    "N_terms",
    "optimizer_method",
    "success_boundary",
    "success_steady",
    "success_transient",
    "boundary_temp_grad",
    "boundary_alpha",
    "boundary_tau",
    "boundary_rms",
    "boundary_temp_grad_95",
    "boundary_alpha_95",
    "boundary_tau_95",
    "steady_temp_grad",
    "steady_h_over_k",
    "steady_rms",
    "steady_temp_grad_95",
    "steady_h_over_k_95",
    "h_initial",
    "k_initial",
    "Pss_initial",
    "h",
    "k",
    "Pss",
    "transient_rms",
    "h_95",
    "k_95",
    "Pss_95",
    "duty_cycle_initial",
    "duty_cycle_final",
    "duty_cycle_mean",
    "plot_file",
    "error_message",
]


def natural_convection_h_guess(v: float, diameter: float) -> float:
    """Initial h guess from the Churchill-Bernstein-style cylinder correlation used in the original script."""
    rho_air = 1.23
    mu_air = 1.789e-5
    pr = 0.71
    k_air = 0.02602

    re = rho_air * v * diameter / mu_air
    return k_air / diameter * (
        0.3
        + 0.62 * re**0.5 * pr ** (1.0 / 3.0)
        / (1.0 + (0.4 / pr) ** (2.0 / 3.0)) ** 0.25
        * (1.0 + (re / 282000.0) ** (5.0 / 8.0)) ** (-4.0 / 5.0)
    )


@dataclass(frozen=True)
class LoadedData:
    """Parsed acquisition data used by the estimator."""

    t: np.ndarray
    tc: np.ndarray
    duty_cycle: np.ndarray
    n_columns_detected: int
    skipped_preamble_rows: int


def _parse_numeric_row(row: Sequence[str], min_cols: int = 8) -> tuple[list[float], int] | None:
    """Return all numeric values from a data row if the first min_cols are numeric.

    Only the first eight columns are currently used by the estimator, but the
    parser accepts and counts additional numeric acquisition channels so files
    with 21 columns do not fail just because they contain extra data.
    """
    if len(row) < min_cols:
        return None

    stripped = [str(cell).strip() for cell in row]
    try:
        first = [float(cell) for cell in stripped[:min_cols]]
    except ValueError:
        return None

    # Count additional numeric columns for bookkeeping. Stop at the first
    # nonnumeric trailing value because some loggers may append comments.
    values = first
    for cell in stripped[min_cols:]:
        if cell == "":
            break
        try:
            values.append(float(cell))
        except ValueError:
            break
    return values, len(values)


def load_six_tc_csv(path: Path) -> LoadedData:
    """Return parsed time, TC, duty-cycle, and file-format bookkeeping.

    The acquisition files may contain metadata lines such as program name,
    acquisition code version, firmware version, date/time stamp, operator notes,
    or a column-name header before the tabular data. This reader scans until it
    finds rows whose first eight columns are numeric:

        time, TC1, TC2, TC3, TC4, TC5, TC6, duty_cycle

    Extra numeric columns are accepted and ignored by the current estimator.
    """
    rows: list[list[float]] = []
    skipped_preamble = 0
    max_columns_detected = 0

    with path.open("r", newline="", errors="ignore") as fp:
        reader = csv.reader(fp)
        for row in reader:
            if not row or all(not str(cell).strip() for cell in row):
                skipped_preamble += 1
                continue

            parsed = _parse_numeric_row(row, 8)
            if parsed is None:
                skipped_preamble += 1
                continue

            values, n_cols = parsed
            max_columns_detected = max(max_columns_detected, n_cols)
            rows.append(values[:8])

    if not rows:
        raise ValueError(
            f"{path.name} has no numeric data rows. Expected at least eight numeric columns: "
            "time, TC1, TC2, TC3, TC4, TC5, TC6, duty_cycle"
        )

    data = np.asarray(rows, dtype=float)

    t = np.asarray(data[:, 0], dtype=float)
    tt = np.asarray(data[:, 1:7], dtype=float).T
    duty_cycle = np.asarray(data[:, 7], dtype=float)

    good = np.isfinite(t) & np.isfinite(duty_cycle)
    for row in tt:
        good &= np.isfinite(row)

    t = t[good]
    tt = tt[:, good]
    duty_cycle = duty_cycle[good]

    if t.size < 10:
        raise ValueError(f"{path.name} has too few valid rows after parsing")

    return LoadedData(
        t=t,
        tc=tt,
        duty_cycle=duty_cycle,
        n_columns_detected=max_columns_detected,
        skipped_preamble_rows=skipped_preamble,
    )


def rms_and_covariance(residuals: np.ndarray, jacobian: np.ndarray, n_params: int) -> tuple[float, np.ndarray]:
    """
    Return residual RMS estimate and 95% parameter half-widths.

    The old script used sqrt(sum(abs(residual))/(dof)), which is not the usual
    least-squares variance estimate. Here we use sqrt(RSS/dof) and the diagonal
    of inv(J'J). A pseudo-inverse is used if J'J is singular or nearly singular.
    """
    residuals = np.asarray(residuals, dtype=float)
    jacobian = np.asarray(jacobian, dtype=float)
    dof = max(residuals.size - n_params, 1)
    sigma = float(np.sqrt(np.sum(residuals**2) / dof))

    jtj = jacobian.T @ jacobian
    try:
        cov = np.linalg.inv(jtj)
    except np.linalg.LinAlgError:
        cov = np.linalg.pinv(jtj)

    diag = np.diag(cov)
    half_widths = 1.96 * sigma * np.sqrt(np.where(diag >= 0.0, diag, np.nan))
    return sigma, half_widths


def fit_least_squares(fun, x0: Sequence[float], method: str, bounds=None, max_nfev: int | None = None):
    kwargs = dict(method=method, max_nfev=max_nfev, x_scale="jac")
    if method != "lm" and bounds is not None:
        kwargs["bounds"] = bounds
    return least_squares(fun, x0, **kwargs)


def fit_boundary_model(t: np.ndarray, tc: np.ndarray, x_tc: Sequence[float], method: str, max_nfev: int | None):
    x1 = float(x_tc[1])
    boundary_gradient = (3.0 * tc[0] - 4.0 * tc[1] + tc[2]) / (2.0 * x1)

    def residual(v_scaled: np.ndarray) -> np.ndarray:
        temp_grad = v_scaled[0]
        alpha = v_scaled[1] / 10000.0
        tau = v_scaled[2] / 10.0
        pred = temp_grad * (1.0 - np.exp(-alpha * (t + tau)))
        if not np.all(np.isfinite(pred)):
            return np.full_like(t, 1.0e12)
        return pred - boundary_gradient

    x0 = np.array([3000.0, 700.0, 30.0], dtype=float)
    # temp_grad >= 0, alpha >= 0, tau may be slightly negative due to trigger/timing offset.
    bounds = ([0.0, 1.0e-12, -1.0e5], [np.inf, np.inf, 1.0e5])
    out = fit_least_squares(residual, x0, method=method, bounds=bounds, max_nfev=max_nfev)

    rms, hw_scaled = rms_and_covariance(out.fun, out.jac, 3)
    values = np.array([out.x[0], out.x[1] / 10000.0, out.x[2] / 10.0], dtype=float)
    half_widths = np.array([hw_scaled[0], hw_scaled[1] / 10000.0, hw_scaled[2] / 10.0], dtype=float)
    return out, values, rms, half_widths, boundary_gradient


def fit_steady_state_model(tc: np.ndarray, x_tc_repeated: np.ndarray, material: MaterialCase, method: str, max_nfev: int | None):
    n_tail = min(20, tc.shape[1])
    tss = tc[:, -n_tail:].reshape(-1)
    xss = x_tc_repeated[:, -n_tail:].reshape(-1)
    diameter = material.diameter
    length = material.length

    def residual(v: np.ndarray) -> np.ndarray:
        if v[1] <= 0:
            return np.full_like(tss, 1.0e12)
        m = np.sqrt(4.0 * v[1] / diameter)
        pred = v[0] / m * np.cosh(m * (length - xss)) / np.sinh(m * length)
        if not np.all(np.isfinite(pred)):
            return np.full_like(tss, 1.0e12)
        return pred - tss

    x0 = np.array([50.0, 1.0], dtype=float)
    bounds = ([0.0, 1.0e-12], [np.inf, np.inf])
    out = fit_least_squares(residual, x0, method=method, bounds=bounds, max_nfev=max_nfev)
    rms, hw = rms_and_covariance(out.fun, out.jac, 2)
    return out, out.x.copy(), rms, hw


class TransientModel:
    """Fourier-series transient model for one material/file."""

    def __init__(
        self,
        material: MaterialCase,
        n_terms: int,
        alpha: float,
        tau: float,
        t_vec: np.ndarray,
        x_vec: np.ndarray,
    ) -> None:
        self.material = material
        self.n_terms = int(n_terms)
        self.alpha = float(alpha)
        self.tau = float(tau)
        self.t = np.asarray(t_vec, dtype=float)
        self.x = np.asarray(x_vec, dtype=float)

        self.area = math.pi * material.diameter**2 / 4.0
        self.perimeter = math.pi * material.diameter
        self.n = np.arange(1, self.n_terms + 1, dtype=float)[:, None]
        self.lambda2 = (self.n * math.pi / material.length) ** 2  # shape (N, 1)
        self.cos_basis = np.cos((self.n * math.pi / material.length) * self.x[None, :])

    def predict(self, params: Sequence[float]) -> np.ndarray:
        h, k, pss = [float(v) for v in params]
        if h <= 0.0 or k <= 0.0 or not np.isfinite(h + k + pss):
            return np.full_like(self.t, np.nan, dtype=float)

        mat = self.material
        a = self.area
        s = self.perimeter
        rho = mat.rho
        c = mat.c_p
        alpha = self.alpha
        tau = self.tau

        nu = h * s / (rho * a * c)
        kappa = k / (rho * c)  # thermal diffusivity K in the original script
        beta2 = kappa * self.lambda2

        t = self.t[None, :]

        # The following algebra intentionally mirrors the original expression.
        with np.errstate(over="ignore", invalid="ignore", divide="ignore"):
            numer = (
                (-beta2 - nu) * np.exp((-beta2 - nu) * t - alpha * tau)
                + (beta2 - alpha + nu) * np.exp(-t * (beta2 + nu))
                + (beta2 + nu) * np.exp(-alpha * (t + tau))
                - beta2
                + alpha
                - nu
            )
            denom = k * mat.length * a * (-beta2 + alpha - nu) * (beta2 + nu)
            modal_coeff = 2.0 * kappa * pss * numer / denom
            modal_sum = np.sum(modal_coeff * self.cos_basis, axis=0)

            base = -(
                np.exp(-alpha * tau - nu * self.t) * nu
                - np.exp(-alpha * (self.t + tau)) * nu
                + (alpha - nu) * (np.exp(-nu * self.t) - 1.0)
            ) * pss * kappa / (k * mat.length * a * nu * (alpha - nu))

            pred = base + modal_sum

        return pred

    def residual(self, params: Sequence[float], observed: np.ndarray) -> np.ndarray:
        pred = self.predict(params)
        if not np.all(np.isfinite(pred)):
            return np.full_like(observed, 1.0e12, dtype=float)
        return pred - observed


def fit_transient_model(
    model: TransientModel,
    observed: np.ndarray,
    h0: float,
    k0: float,
    pguess: float,
    method: str,
    max_nfev: int | None,
):
    x0 = np.array([h0, k0, pguess], dtype=float)
    x0 = np.where(np.isfinite(x0), x0, np.array([10.0, 10.0, 1.0]))
    x0 = np.maximum(x0, np.array([1.0e-9, 1.0e-9, 1.0e-12]))

    def residual(v: np.ndarray) -> np.ndarray:
        return model.residual(v, observed)

    bounds = ([1.0e-12, 1.0e-12, 1.0e-12], [np.inf, np.inf, np.inf])
    out = fit_least_squares(residual, x0, method=method, bounds=bounds, max_nfev=max_nfev)
    rms, hw = rms_and_covariance(out.fun, out.jac, 3)
    return out, out.x.copy(), rms, hw


def select_files(material: MaterialCase, test: str, explicit_files: Sequence[str] | None):
    if explicit_files:
        # Accept names with or without .csv.
        files = tuple(Path(f).stem for f in explicit_files)
        v_guess = (5.0,) * len(files)
        ymax = (120.0,) * len(files)
        tinf = material.temp_tinf if test == "temp" else material.wind_tinf
        return files, v_guess, ymax, tinf

    if test == "temp":
        return material.temp_files, material.temp_v_guess, material.temp_ymax, material.temp_tinf
    return material.wind_files, material.wind_v_guess, material.wind_ymax, material.wind_tinf


def infer_material_key_from_filename(stem: str) -> str | None:
    """Infer material from a filename stem such as 50al1, 75cu1, or 100ss1."""
    lower = stem.lower()
    # Check ss before the single-letter-ish cases to avoid accidental matches.
    if "ss" in lower:
        return "ss"
    if "cu" in lower:
        return "cu"
    if "al" in lower:
        return "al"
    return None


def default_ymax_for_file(stem: str, test: str, material: MaterialCase) -> float:
    lower = stem.lower()
    if test == "temp":
        if lower.startswith("50"):
            return 60.0
        if lower.startswith("75"):
            return 90.0
        if lower.startswith("100"):
            return 120.0
    if test == "wind":
        return 90.0
    return 120.0


def build_run_tasks(args: argparse.Namespace, data_dir: Path, plots_dir: Path) -> tuple[list[dict[str, object]], list[Path]]:
    """Build independent per-file work items for serial or parallel execution.

    Default behavior is directory discovery: scan DATA_DIR for CSV files whose
    names identify one of the known materials. This avoids noisy reports about
    files in the historical built-in lists that are not present in a particular
    working directory.

    Use --expected-list to reproduce the old behavior of selecting from the
    built-in temperature/wind filename lists. Use --files to force specific
    basenames or CSV paths.
    """
    material_keys = tuple(MATERIALS.keys()) if args.material == "all" else (args.material,)
    tasks: list[dict[str, object]] = []
    missing: list[Path] = []

    if args.files and args.expected_list:
        raise ValueError("Use either --files or --expected-list, not both.")

    # Discovery is the default. The --discover flag is retained as a harmless
    # compatibility alias for scripts written against v2.2.x.
    discovery_mode = bool(args.discover) or (not args.expected_list and not args.files)

    if discovery_mode:
        csv_paths = sorted(data_dir.glob("*.csv"))
        for path in csv_paths:
            stem = path.stem
            inferred_key = infer_material_key_from_filename(stem)
            if inferred_key is None:
                # Unknown CSVs are ignored silently in discovery mode. To run an
                # unlabeled file, provide it explicitly with --material ... --files ...
                continue
            if inferred_key not in material_keys:
                continue

            mat = MATERIALS[inferred_key]
            tinf = mat.temp_tinf if args.test == "temp" else mat.wind_tinf
            tasks.append(dict(
                task_index=len(tasks),
                material=mat,
                test=args.test,
                filename=stem,
                v_guess=5.0,
                y_max=default_ymax_for_file(stem, args.test, mat),
                tinf=tinf,
                data_dir=data_dir,
                plots_dir=plots_dir,
                n_terms=args.N,
                method=args.method,
                max_nfev=args.max_nfev,
                show=args.show,
                dpi=args.dpi,
                include_title=args.title,
            ))
        return tasks, missing

    for key in material_keys:
        mat = MATERIALS[key]
        files, v_guesses, y_maxes, tinf = select_files(mat, args.test, args.files)

        for idx, filename in enumerate(files):
            stem = Path(filename).stem
            in_path = data_dir / f"{stem}.csv"
            if not in_path.exists():
                missing.append(in_path)
                if args.strict:
                    raise FileNotFoundError(f"[missing] {in_path}")
                # Keep default output quiet. Missing expected-list files can be
                # reported at the end with --report-missing.
                continue

            tasks.append(dict(
                task_index=len(tasks),
                material=mat,
                test=args.test,
                filename=stem,
                v_guess=v_guesses[min(idx, len(v_guesses) - 1)],
                y_max=y_maxes[min(idx, len(y_maxes) - 1)],
                tinf=tinf,
                data_dir=data_dir,
                plots_dir=plots_dir,
                n_terms=args.N,
                method=args.method,
                max_nfev=args.max_nfev,
                show=args.show,
                dpi=args.dpi,
                include_title=args.title,
            ))

    return tasks, missing

def _run_one_file_from_task(task: dict[str, object]) -> dict[str, object]:
    """Small top-level wrapper so ProcessPoolExecutor can pickle the work item."""
    task = dict(task)
    task.pop("task_index", None)
    return run_one_file(**task)


def run_tasks(tasks: list[dict[str, object]], *, jobs: int, strict: bool) -> list[dict[str, object]]:
    """Run per-file tasks serially or in parallel, preserving input order in results."""
    if not tasks:
        return []

    rows_by_index: dict[int, dict[str, object]] = {}

    if jobs <= 1 or len(tasks) == 1:
        for task in tasks:
            idx = int(task["task_index"])
            mat = task["material"]
            print(f"\nProcessing {mat.label} / {task['test']} / {task['filename']}.csv")
            try:
                row = _run_one_file_from_task(task)
            except Exception as exc:
                if strict:
                    raise
                print(f"  [failed] {task['filename']}.csv: {exc}")
                row = {
                    "version": VERSION,
                    "run_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "material": mat.key,
                    "test": task["test"],
                    "filename": task["filename"],
                    "error_message": str(exc),
                }
            rows_by_index[idx] = row
            if not row.get("error_message"):
                print("  h={h:.6g}, k={k:.6g}, Pss={Pss:.6g}, transient RMS={rms:.6g}".format(
                    h=row["h"], k=row["k"], Pss=row["Pss"], rms=row["transient_rms"]
                ))
        return [rows_by_index[i] for i in sorted(rows_by_index)]

    print(f"\nRunning {len(tasks)} file(s) with {jobs} worker process(es).")
    with ProcessPoolExecutor(max_workers=jobs) as executor:
        future_to_task = {executor.submit(_run_one_file_from_task, task): task for task in tasks}
        for future in as_completed(future_to_task):
            task = future_to_task[future]
            idx = int(task["task_index"])
            mat = task["material"]
            try:
                row = future.result()
            except Exception as exc:
                if strict:
                    raise
                print(f"  [failed] {mat.label} / {task['filename']}.csv: {exc}")
                row = {
                    "version": VERSION,
                    "run_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "material": mat.key,
                    "test": task["test"],
                    "filename": task["filename"],
                    "error_message": str(exc),
                }
            rows_by_index[idx] = row
            if not row.get("error_message"):
                print("  [done] {label} / {file}.csv: h={h:.6g}, k={k:.6g}, Pss={Pss:.6g}, RMS={rms:.6g}".format(
                    label=mat.label, file=task["filename"], h=row["h"], k=row["k"], Pss=row["Pss"], rms=row["transient_rms"]
                ))

    return [rows_by_index[i] for i in sorted(rows_by_index)]



def make_x_tc_matrix(x_tc: Sequence[float], n: int) -> np.ndarray:
    return np.tile(np.asarray(x_tc, dtype=float)[:, None], (1, n))


def plot_fit(
    path: Path,
    t: np.ndarray,
    tc: np.ndarray,
    material: MaterialCase,
    tinf: float,
    y_max: float,
    ptime: float,
    model: TransientModel,
    params: Sequence[float],
    title: str | None,
    show: bool,
    dpi: int,
) -> None:
    try:
        plt.rcParams.update({"font.family": "serif", "font.serif": ["Times New Roman", "Times", "DejaVu Serif"]})
    except Exception:
        pass

    fig, ax = plt.subplots(figsize=(8.0, 5.0))

    for i in range(6):
        xrow = np.full_like(t, material.x_tc[i], dtype=float)
        local_model = TransientModel(material, model.n_terms, model.alpha, model.tau, t, xrow)
        pred = local_model.predict(params)

        label_data = "Experimental data" if i == 0 else None
        label_model = "Model" if i == 0 else None
        ax.plot(t, tc[i] + tinf, "k.", markersize=3, label=label_data)
        ax.plot(t, pred + tinf, "-r", linewidth=2, label=label_model)

        label_x = min(float(ptime), float(np.nanmax(t)))
        y_last = float(tc[i, -1] + tinf - 4.0)
        ax.text(label_x, y_last, f"TC {i + 1}", va="bottom", ha="right", fontsize=11)

    ax.set_xlim(0.0, ptime)
    ax.set_ylim(20.0, y_max)
    ax.set_xlabel("Time (s)")
    ax.set_ylabel(r"Temperature ($^\circ$C)")
    if title:
        ax.set_title(title)
    ax.grid(False)
    ax.legend(loc="upper left", frameon=False)
    fig.tight_layout()
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, format="pdf", dpi=dpi)

    if show:
        plt.show()
    else:
        plt.close(fig)


def write_results_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=RESULT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_results_xlsx(path: Path, rows: list[dict[str, object]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(RESULT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in RESULT_COLUMNS])

    for col_idx, name in enumerate(RESULT_COLUMNS, start=1):
        width = min(max(len(name) + 2, 12), 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def run_one_file(
    *,
    material: MaterialCase,
    test: str,
    filename: str,
    v_guess: float,
    y_max: float,
    tinf: float,
    data_dir: Path,
    plots_dir: Path,
    n_terms: int,
    method: str,
    max_nfev: int | None,
    show: bool,
    dpi: int,
    include_title: bool,
) -> dict[str, object]:
    in_path = data_dir / f"{filename}.csv"
    loaded = load_six_tc_csv(in_path)
    t = loaded.t
    tc = loaded.tc
    duty_cycle = loaded.duty_cycle
    x_tc_matrix = make_x_tc_matrix(material.x_tc, t.size)

    # Flatten in the same TC-block order as the original script:
    # all TC1 times, then all TC2 times, ..., TC6.
    t_stacked = np.tile(t, 6)
    x_stacked = x_tc_matrix.reshape(-1)
    temp_stacked = tc.reshape(-1)

    boundary_out, boundary_vals, boundary_rms, boundary_hw, _boundary_gradient = fit_boundary_model(
        t, tc, material.x_tc, method, max_nfev
    )
    alpha = float(boundary_vals[1])
    tau = float(boundary_vals[2])

    steady_out, steady_vals, steady_rms, steady_hw = fit_steady_state_model(
        tc, x_tc_matrix, material, method, max_nfev
    )

    h0 = natural_convection_h_guess(v_guess, material.diameter)
    h_over_k = float(steady_vals[1])
    k0 = h0 / h_over_k if h_over_k > 0.0 else np.nan
    pguess = float(steady_vals[0]) * k0 * (math.pi * material.diameter**2 / 4.0)

    transient_model = TransientModel(material, n_terms, alpha, tau, t_stacked, x_stacked)
    transient_out, transient_vals, transient_rms, transient_hw = fit_transient_model(
        transient_model, temp_stacked, h0, k0, pguess, method, max_nfev
    )

    plot_path = plots_dir / f"{filename}_fit.pdf"
    title = f"{material.label} {filename}: h={transient_vals[0]:.4g}, k={transient_vals[1]:.4g} W/(m K)" if include_title else None
    plot_fit(
        plot_path,
        t,
        tc,
        material,
        tinf,
        y_max,
        material.plot_time,
        TransientModel(material, n_terms, alpha, tau, t, np.zeros_like(t)),
        transient_vals,
        title,
        show,
        dpi,
    )

    return {
        "version": VERSION,
        "run_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "material": material.key,
        "test": test,
        "filename": filename,
        "n_rows": int(t.size),
        "n_columns_detected": int(loaded.n_columns_detected),
        "skipped_preamble_rows": int(loaded.skipped_preamble_rows),
        "N_terms": int(n_terms),
        "optimizer_method": method,
        "success_boundary": bool(boundary_out.success),
        "success_steady": bool(steady_out.success),
        "success_transient": bool(transient_out.success),
        "boundary_temp_grad": boundary_vals[0],
        "boundary_alpha": boundary_vals[1],
        "boundary_tau": boundary_vals[2],
        "boundary_rms": boundary_rms,
        "boundary_temp_grad_95": boundary_hw[0],
        "boundary_alpha_95": boundary_hw[1],
        "boundary_tau_95": boundary_hw[2],
        "steady_temp_grad": steady_vals[0],
        "steady_h_over_k": steady_vals[1],
        "steady_rms": steady_rms,
        "steady_temp_grad_95": steady_hw[0],
        "steady_h_over_k_95": steady_hw[1],
        "h_initial": h0,
        "k_initial": k0,
        "Pss_initial": pguess,
        "h": transient_vals[0],
        "k": transient_vals[1],
        "Pss": transient_vals[2],
        "transient_rms": transient_rms,
        "h_95": transient_hw[0],
        "k_95": transient_hw[1],
        "Pss_95": transient_hw[2],
        "duty_cycle_initial": float(duty_cycle[0]),
        "duty_cycle_final": float(duty_cycle[-1]),
        "duty_cycle_mean": float(np.mean(duty_cycle)),
        "plot_file": str(plot_path),
        "error_message": "",
    }


def existing_data_dir_near_script() -> Path:
    script_dir = Path(__file__).resolve().parent
    candidate = script_dir / "data"
    return candidate if candidate.exists() else script_dir


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Estimate h, k, and steady power for 1-D transient convection data.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    
    parser.add_argument(
        "--version",
        action="version",
        version=f"{PROGRAM_NAME} version {VERSION}",
        help="show program version and exit",
    )
    
    parser.add_argument("--material", choices=("al", "cu", "ss", "all"), default="al", help="material case to run")
    parser.add_argument("--test", choices=("temp", "wind"), default="temp", help="test set to run")
    parser.add_argument("--files", nargs="+", help="specific CSV basenames or filenames to process")
    parser.add_argument("--discover", action="store_true", help="scan DATA_DIR for CSV files and infer material from filenames; this is already the default unless --files or --expected-list is used")
    parser.add_argument("--expected-list", action="store_true", help="use the built-in historical file lists for the selected material/test instead of scanning DATA_DIR")
    parser.add_argument("--report-missing", action="store_true", help="when using --expected-list, print the expected CSV files that were not found")
    parser.add_argument("--data-dir", type=Path, default=existing_data_dir_near_script(), help="directory containing CSV data")
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parent, help="directory for result files")
    parser.add_argument("--plots-dir", type=Path, help="directory for PDF plots; default is OUTPUT_DIR/FIGS")
    parser.add_argument("--results-name", default="oneDkh_results", help="base filename for CSV/XLSX results")
    parser.add_argument("--N", type=int, default=100, help="number of Fourier terms")
    parser.add_argument("--method", choices=("trf", "lm"), default="trf", help="least_squares method")
    parser.add_argument("--max-nfev", type=int, default=None, help="maximum least-squares function evaluations")
    parser.add_argument("--strict", action="store_true", help="fail if any selected CSV file is missing")
    parser.add_argument("--show", action="store_true", help="show plots interactively")
    parser.add_argument("--dpi", type=int, default=150, help="plot DPI")
    parser.add_argument("--title", action="store_true", help="include fitted parameter summary in plot title")
    parser.add_argument("--jobs", type=int, default=0, help="parallel worker processes; 0 means all available CPU cores, 1 means serial")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)

    data_dir = args.data_dir.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    plots_dir = (args.plots_dir.expanduser().resolve() if args.plots_dir else output_dir / "FIGS")
    output_dir.mkdir(parents=True, exist_ok=True)
    plots_dir.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, object]] = []

    if args.N < 1:
        raise ValueError("--N must be at least 1")

    tasks, missing = build_run_tasks(args, data_dir, plots_dir)

    cpu_count = os.cpu_count() or 1
    requested_jobs = cpu_count if args.jobs == 0 else max(1, args.jobs)
    jobs = min(requested_jobs, max(1, len(tasks)))
    if args.show and jobs > 1:
        print("--show requested; forcing --jobs 1 because interactive plotting and multiprocessing do not mix well.")
        jobs = 1

    selection_mode = "explicit file list" if args.files else ("built-in expected list" if args.expected_list else "data-directory discovery")

    print(f"oneDkhEstimator.py v{VERSION}")
    print(f"Data directory:   {data_dir}")
    print(f"Output directory: {output_dir}")
    print(f"Plots directory:  {plots_dir}")
    print(f"Selected files:   {len(tasks)} CSV file(s) ({selection_mode})")
    print(f"Worker processes: {jobs}")
    if requested_jobs > jobs and len(tasks) > 0:
        print(f"Worker cap:       requested {requested_jobs}, capped at {jobs} because only {len(tasks)} file(s) are selected")

    rows = run_tasks(tasks, jobs=jobs, strict=args.strict)

    if not rows:
        print("\nNo files were processed.")
        if missing:
            print("Missing files:")
            for path in missing:
                print(f"  {path}")
        return 1

    csv_path = output_dir / f"{args.results_name}.csv"
    xlsx_path = output_dir / f"{args.results_name}.xlsx"
    write_results_csv(csv_path, rows)
    wrote_xlsx = write_results_xlsx(xlsx_path, rows)

    print(f"\nProcessed {len(rows)} file(s).")
    print(f"Wrote CSV:  {csv_path}")
    if wrote_xlsx:
        print(f"Wrote XLSX: {xlsx_path}")
    else:
        print("openpyxl is not installed; skipped XLSX output.")

    if missing and args.report_missing and not args.strict:
        print("\nExpected-list files not found:")
        for path in missing:
            print(f"  {path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
