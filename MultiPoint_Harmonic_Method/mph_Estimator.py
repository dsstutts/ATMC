# This code uses the model outlined in Tomanek, Lauren B. and Stutts, Daniel S.
# "Thermal conductivity estimation via a multi-point harmonic one-dimensional
# convection model," 2022
# International Journal of Heat and Mass Transfer , Vol. 186
# Elsevier
# pp. 1-6 No. 122467.
#
# To run the code. Set the file directiories and filename for results and material to test,
# and then execute the code.  The input file data may be either tab or comma
# separated.
#
# The output consists of a plots of the boundary input model and overall harmonic
# model with using the optimal parameters found based on the input data,
# and an excel document containing the relevant results.
#
# The estimation is performed by using the scipy.optimize.least_squares function. For
# each estimation, the optimal parameters, mean standard error, and individual
# parameter uncertainties are calculated.
#
# This code is copyrighted by the authors, but released under the MIT
# license:
#
# Copyright (c) 2025 -- mph_Estimator.py
#
# S&T and the University of Missouri Board of Curators
# license to you the right to use, modify, copy, and distribute this
# code subject to the MIT license:
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included
# in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
# FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.
###################################################

import sys
from scipy.optimize import least_squares
import numpy as np
import scipy.stats
from numpy import array
import re

# from xlwt import Workbook # Writing to an excel
import array as arr
from pathlib import Path
import matplotlib.pyplot as plt
import matplotlib as mpl

mpl.rcParams["text.usetex"] = True
mpl.rcParams["font.family"] = "serif"
import matplotlib as mpl

mpl.rcParams.update(
    {
        "text.usetex": False,  # Matplotlib draws text
        "font.family": "serif",
        "font.serif": ["Times New Roman"],  # or "Times"
        "mathtext.fontset": "stix",  # Times-like math glyphs
        "pdf.fonttype": 42,  # embed TrueType in PDF
        "ps.fonttype": 42,  # if you also save EPS
        "svg.fonttype": "none",  # keep text as text (if you export SVG)
    }
)

from openpyxl import Workbook

wb = Workbook()
# rename the default sheet (or create a new one)
sheet1 = wb.active
# plottype = ''# Defaults to PNG
plottype = "PDF"
# Set the desired resolution:
PLOT_DPI = 120  # Dpi.  Fine for PDF, but should use higher for PNG.
cur_version = sys.version_info  # Test for Python version:
BASE = Path(__file__).resolve().parent  # Set base directory

# locations
filefolder = BASE / "data"
excelsaveloc = BASE
plotsaveloc = BASE / "FIGS"
Tinf = 22

period = 300  # oscillation period
maxloc = period / 2
# Wind = True
Temp = True
mm = 1  # Set index to pick different windspeeds if necessary

# Select material:

Al = False
SS = True
Cu = False


def plot_lbc_model_vs_data(
    t,
    data_y,
    theta_A,
    theta_M,
    tau,
    omega,
    out_dir: Path,
    fname_base: str,
    dpi: int = 500,
):
    """
    Plot fitted cosine model vs data and save as PNG and PDF
    in both out_dir and out_dir/FIGS.

    Parameters
    ----------
    t : array-like
        Time vector (seconds).
    data_y : array-like
        Measured data to fit (same length as t).
    theta_A, theta_M, tau : float
        Fitted parameters: amplitude, mean (offset), time shift.
    omega : float
        Angular frequency (rad/s).
    out_dir : Path
        Base directory for outputs.
    fname_base : str
        Base filename (no extension).
    dpi : int
        Raster resolution for PNG (safe to pass to PDF too).
    """
    t_arr = np.asarray(t, dtype=float)
    y_data = np.asarray(data_y, dtype=float)
    y_fit = theta_A * np.cos(omega * (t_arr - tau)) + theta_M

    # Ensure directories exist
    out_dir = Path(out_dir)
    figs_dir = out_dir / "FIGS"
    # out_dir.mkdir(parents=True, exist_ok=True)
    figs_dir.mkdir(parents=True, exist_ok=True)

    # Figure
    fig, ax = plt.subplots()
    ax.plot(t_arr, y_fit, "-", linewidth=2, color="red", label="Model")
    ax.plot(t_arr, y_data, "o", markersize=4, mfc="green", mec="none", label="Data")
    ax.set_xlabel("t (s)")
    ax.set_ylabel(r"$\theta_{1}$ (°C)")
    ax.set_title(
        r"Fit: $\theta_{1}(t)=A\cos(\omega\,(t-\tau))+M$"
        + "\n"
        + rf"$A={theta_A:.3g},\;M={theta_M:.3g},\;\tau={tau:.3g},\;\omega={omega:.3g}$"
    )
    ax.legend()
    ax.grid(True, alpha=0.3)

    base = f"{fname_base}_fit"

    # # Save in out_dir
    # png_main = out_dir / f"{base}.png"
    # pdf_main = out_dir / f"{base}.pdf"
    # fig.savefig(png_main, format="png", dpi=dpi, bbox_inches="tight")
    # fig.savefig(pdf_main, format="pdf", dpi=dpi, bbox_inches="tight")

    # Save in out_dir/FIGS
    png_figs = figs_dir / f"{base}.png"
    pdf_figs = figs_dir / f"{base}.pdf"
    fig.savefig(png_figs, format="png", dpi=dpi, bbox_inches="tight")
    fig.savefig(pdf_figs, format="pdf", dpi=dpi, bbox_inches="tight")

    plt.close(fig)

    return {
        "png_figs": png_figs,
        "pdf_figs": pdf_figs,
    }


# Plot harmonic model and data


def plot_all_x_overlay(
    tt,
    x,
    T,
    model_fn,
    V,
    *,
    TTT=None,  # optional passthrough to model_fn
    XX_preferred=None,  # order legend/trace
    emphasize_first_x=True,
    figsize=(10, 6),
    show=True,
    plottype="PNG",  # "PNG" or "PDF"
    dpi=300,  # used for raster formats; safe to keep for PDF too
    out_dir=Path("."),
    fname_base="harmonic_model_and_data",
    model_kwargs=None,
    atol=1e-12,
):
    """
    Single-axes overlay: plot all x-location temperatures vs time.
    Data -> green circles; Model -> red lines.

    Parameters
    ----------
    tt, x, T : 1D arrays
        Flattened time, position, and temperature vectors (like least_squares inputs).
    model_fn : callable
        model_fn(V, t_vec, x_scalar, TTT?, **model_kwargs) -> array(len(t_vec))
        If TTT is None or model_fn doesn't accept it, it is omitted.
    V : sequence
        Model parameters (e.g., [h, k]).
    XX_preferred : array-like | None
        Desired order of x-traces in the legend; only those present in `x` are kept.
    plottype : {"PNG","PDF"}
        Output format. PDF is best for publication; PNG for slides/docs.
    dpi : int
        Raster resolution (also affects some text metrics during layout).
    out_dir : Path
    fname_base : str
    Returns
    -------
    plot_path : Path
        The filepath that was written.
    """

    model_kwargs = {} if model_kwargs is None else dict(model_kwargs)

    tt = np.asarray(tt, float).ravel()
    xx = np.asarray(x, float).ravel()
    YY = np.asarray(T, float).ravel()

    # de-duped x list (preserves user-preferred order)
    if XX_preferred is not None:
        pref = np.asarray(XX_preferred, float).ravel()
        x_list = [
            float(v) for v in pref if np.any(np.isclose(xx, v, rtol=0, atol=atol))
        ]
    else:
        x_list = []
        for v in xx:
            if not any(np.isclose(v, s, rtol=0, atol=atol) for s in x_list):
                x_list.append(float(v))

    fig, ax = plt.subplots(figsize=figsize)

    y_min, y_max = np.inf, -np.inf

    for j, xv in enumerate(x_list):
        mask = np.isclose(xx, xv, rtol=0, atol=atol)
        tj = tt[mask]
        yj = YY[mask]
        if tj.size == 0:
            continue

        order = np.argsort(tj)
        tj, yj = tj[order], yj[order]

        # data: green filled circles
        ax.scatter(tj, yj, s=20, color="green", marker="o", zorder=3, label=None)

        # model: red line (evaluate at the same times)
        try:
            y_model = model_fn(V, tj, float(xv), TTT, **model_kwargs)
        except TypeError:
            # model_fn doesn't accept TTT
            y_model = model_fn(V, tj, float(xv), **model_kwargs)

        y_model = np.asarray(y_model).ravel()
        if y_model.shape[0] != tj.shape[0]:
            raise ValueError(
                "Model output length must match the time vector for scalar x."
            )

        style = (
            dict(color="red", linewidth=2.2)
            if (emphasize_first_x and j == 0)
            else dict(color="red", alpha=0.85, linewidth=1.5)
        )
        ax.plot(tj, y_model, label=f"x={xv:g}", **style)

        # limits
        y_min = min(y_min, np.nanmin(yj), np.nanmin(y_model))
        y_max = max(y_max, np.nanmax(yj), np.nanmax(y_model))

    # Titles/labels: keep mathtext-safe (no \big or \!)
    ax.set_title(r"All $x$ locations: model (red) vs data (green)")
    ax.set_xlabel(r"time, $t$")
    ax.set_ylabel("temperature")
    ax.grid(True, linestyle=":", linewidth=0.8)

    if emphasize_first_x:
        ax.legend(frameon=False, fontsize="small", ncols=2)

    if not np.isfinite(y_min) or not np.isfinite(y_max) or np.isclose(y_min, y_max):
        pad = 0.05 * (abs(y_min) if np.isfinite(y_min) else 1.0) + 1e-9
        y_min = (y_min if np.isfinite(y_min) else -1.0) - pad
        y_max = (y_max if np.isfinite(y_max) else 1.0) + pad
    ax.set_ylim(y_min, y_max)

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if plottype.upper() == "PNG":
        plot_path = out_dir / f"{fname_base}.png"
        fig.savefig(plot_path, format="png", dpi=dpi, bbox_inches="tight")
    elif plottype.upper() == "PDF":
        plot_path = out_dir / f"{fname_base}.pdf"
        fig.savefig(plot_path, format="pdf", dpi=dpi, bbox_inches="tight")
    else:
        raise ValueError("plottype must be 'PNG' or 'PDF'")

    if show:
        plt.show()
    plt.close(fig)

    return plot_path


######### End model and data plotting

if Cu == True:
    filesave = "cu300test_results"  # Excel filename
    sheet1.title = filesave  # worksheet title
    # material properties and files to run
    # copper
    X = [0.000, 0.015, 0.030, 0.045, 0.060, 0.075]  # thermocouple locations in meters
    filename = "cu300test"  # data file
    if Temp == True:
        VV = [5, 5, 5, 5, 5, 5, 5, 5, 5]  # wind speed guess
        ymax = [110, 60, 60, 90, 90, 90, 120, 120, 120]  # yaxis maximum temperature
        Tinf = 22.0
    rho = 8912.93  # density
    D = 0.003175  # rod diameter
    De = 50.8e-6
    L = 0.150  # length
    Le = 0.01e-3
    xe = 0.01e-3
    te = 0.82e-3
    thel = 0.5112
    theh = 0.5468
    c = 384.93  # specific heat
    pT = 600  # plot time seconds

if Al == True:
    # aluminum
    X = [0, 0.01, 0.02, 0.03, 0.04, 0.05]
    if Temp == True:
        filename = "al300test"  # data file
        filesave = "al300test_results"  # Excel filename
        sheet1.title = filesave  # worksheet title
        VV = [5, 5, 5, 5, 5, 5, 5, 5, 5]  # wind speed guess
        ymax = [110, 60, 60, 90, 90, 90, 120, 120, 120]  # yaxis maximum temperature
        Tinf = 22.0
    rho = 2767.99  # density
    D = 0.003175  # rod diameter
    De = 152.4e-6
    L = 0.150  # length
    Le = 0.01e-3
    xe = 0.01e-3
    te = 0.82e-3
    thel = 0.5112
    theh = 0.5468
    c = 896  # specific heat
    pT = 600  # plot time seconds

if SS == True:
    # stainless steel
    filesave = "ss300test_results"  # Excel filename
    sheet1.title = filesave  # worksheet title
    X = [0, 0.007, 0.014, 0.021, 0.028, 0.035]
    if Temp == True:
        filename = "ss300test"  # data file
        VV = [5, 5, 5, 5, 5, 5, 5, 5, 5]  # wind speed guess
        ymax = [110, 60, 60, 90, 90, 90, 120, 120, 120]  # yaxis maximum temperature
        Tinf = 22.0
    rho = 8030  # density
    D = 0.003175  # rod diameter
    De = 25.4e-6
    L = 0.150  # length
    Le = 0.01e-3
    xe = 0.01e-3
    te = 0.82e-3
    thel = 0.5112
    theh = 0.5468
    c = 502  # specific heat
    pT = 600  # plot time seconds

# Define the model:
A = np.pi * D**2 / 4
s = np.pi * D
omega = 2 * np.pi / period
N = 200  # Number of terms in Fourier Series

# Define labels for the first column
labels_col1 = [
    "MP harmonic model",
    "Number of Fourier terms N = ",
    "Boundary temperature model",
    "thetaA",
    "thetaM",
    "tau",
    "Harmonic Model",
    "h",
    "k",
]  # labels for results

# Write number of Fourier terms to the spreadsheet:
sheet1.cell(row=2, column=2, value=N)

# Assign labels to sheet1:

for i in range(1, len(labels_col1) + 1):
    sheet1.cell(row=i, column=1, value=labels_col1[i - 1])

# Define CI lables on row 1 for columns 3 and 4:
labels_row1 = ["95% CI low", "95% CI high"]
# Assign CI column labels:
for j in range(0, len(labels_row1)):
    sheet1.cell(row=1, column=j + 3, value=labels_row1[j])

v = VV[mm]  # windspeed

# Create empty lists:
TT = [[], [], [], [], [], []]
xx = [[], [], [], [], [], []]
t = []
T = []
x = []
Tss = []
xss = []
best_hk = [0.0, 0.0]

infile = filefolder / f"{filename}.csv"
plotname = plotsaveloc / filename  # plot save location
try:
    data = open(infile, "r")  # get array out of input file
except:
    print("Cannot find input file; Please try again.")
    sys.exit(0)

############## Read The Data File ##############

data.seek(0)  # Reset file pointer to the beginning
linecount = 0
# next(data) #skips the header in the file
# Read the data from the input file:
if cur_version[0] == 3:  # This is necesary due to the change in the type
    for line in data:  # returned by the map function in Python 3.x.x.
        linedat = list(
            map(float, re.split("\t|,", line))
        )  # Reads tab and comma-delimited files
        t.append(linedat[0])
        for i in range(0, 6):
            TT[i].append(linedat[i + 1])
        for i in range(6, 12):
            xx[i - 6] = [X[i - 6]] * len(t)
        linecount += 1
else:
    for line in data:  # Python2.xx option (probably won't work anyway)
        linedat = list(map(float, re.split("\t|,", line)))
        t.append(line[0])
        for i in range(0, 6):
            TT[i].append(linedat[i + 1])
        for i in range(6, 12):
            xx[i - 6] = [X[i - 6]] * len(t)
        linecount += 1
# Close the input file:
data.close()

# Arranging data into three column matrices

T = array(TT[1])
x = array(xx[1])
tt = np.concatenate((array(t), array(t), array(t), array(t), array(t)), axis=0)
for i in range(2, 6):
    T = np.concatenate((T, array(TT[i])), axis=0)
for i in range(2, 6):
    x = np.concatenate((x, array(xx[i])), axis=0)

# h initial guess
rhoair = 1.23
muair = 1.789 * 10 ** (-5)
Re = rhoair * v * D / muair
Pr = 0.71
kair = 0.02602
# First estimate for h:
h0 = (
    kair
    / D
    * (
        0.3
        + 0.62
        * Re ** (0.5)
        * Pr ** (1.0 / 3)
        / (1 + (0.4 / Pr) ** (2.0 / 3)) ** (0.25)
        * (1 + (Re / 282000) ** (5.0 / 8)) ** (-4.0 / 5)
    )
)


# Left boundary model
def LBC(V):
    theta__A = V[0]
    theta__M = V[1]
    tau = V[2]
    return theta__A * np.cos(omega * (array(t) - tau)) + theta__M - array(TT[0])


res_LBC = least_squares(LBC, x0=[10, 50, maxloc])

# Best-fit parameters
LBC_params = res_LBC.x
LBC_paramsA = arr.array("d", LBC_params)  # Optimal parameters as array('d', ...)

# Print and write the parameter values

for i in range(0, len(LBC_paramsA)):
    print(labels_col1[i + 3] + " = {}".format(LBC_paramsA[i]))
    sheet1.cell(row=i + 4, column=2, value=LBC_paramsA[i])

# --- 95% confidence intervals from Jacobian ---
# Residuals and problem size
m = res_LBC.fun.size  # number of data points
n = res_LBC.x.size  # number of parameters
dof = max(1, m - n)  # degrees of freedom (guard against <=0)
RMS_LBC = np.sqrt(2 * res_LBC.cost) / dof
J = res_LBC.jac  # Jacobian at the solution (m x n)
JTJ = J.T @ J
print("Harmonic input boundary model degrees of freedom = ", dof)
print("RMS_LBC = ", RMS_LBC)
# Invert or pseudo-invert JTJ robustly
try:
    JTJ_inv = np.linalg.inv(JTJ)
except np.linalg.LinAlgError:
    JTJ_inv = np.linalg.pinv(JTJ)

# Residual variance estimate: cost = 0.5 * RSS
sigma2 = (2.0 * res_LBC.cost) / dof

# Parameter covariance and standard errors
cov_x = sigma2 * JTJ_inv
se = np.sqrt(np.diag(cov_x))

# t critical value for 95% CI
tcrit = scipy.stats.t.ppf(0.975, dof)

# Confidence intervals
ci_LBC_lower = LBC_paramsA - tcrit * se
ci_LBC_upper = LBC_paramsA + tcrit * se

# Print CIs
for j in range(n):
    print(f"{labels_col1[j+3]} 95% CI: [{ci_LBC_lower[j]}, {ci_LBC_upper[j]}]")

sheet1.cell(row=4, column=3, value=ci_LBC_lower[0])
sheet1.cell(row=4, column=4, value=ci_LBC_upper[0])
sheet1.cell(row=5, column=3, value=ci_LBC_lower[1])
sheet1.cell(row=5, column=4, value=ci_LBC_upper[1])
sheet1.cell(row=6, column=3, value=ci_LBC_lower[2])
sheet1.cell(row=6, column=4, value=ci_LBC_upper[2])

theta__A_values = [LBC_paramsA[0]]
theta__M_values = [LBC_paramsA[1]]
tau_values = [LBC_paramsA[2]]

# Plot the model over the data:
paths = plot_lbc_model_vs_data(
    t=t,
    data_y=TT[0],
    theta_A=LBC_paramsA[0],
    theta_M=LBC_paramsA[1],
    tau=LBC_paramsA[2],
    omega=omega,
    out_dir=".",
    fname_base=filename + "harmonic_boundary_input_model",
    dpi=PLOT_DPI,
)
print("Wrote: figures to: ", plotsaveloc)


def harmonic_model_res(V, tT, XX, TTT):
    """
    Residuals for boundary temperature model parameter estimation.
    V[0] = h  (via nu)
    V[1] = k  (via K)
    """
    theta__A = LBC_paramsA[0]
    theta__M = LBC_paramsA[1]
    tau = LBC_paramsA[2]
    h, k = V
    nu = (h * s) / (rho * A * c)  # convective term
    K = k / (rho * c)  # thermal diffusivity

    tT = np.asarray(tT, dtype=float)
    XX = np.asarray(XX, dtype=float)
    TTT = np.asarray(TTT, dtype=float)
    # Create an array of the size of tT to assign model values
    total = np.zeros_like(tT, dtype=float)

    # Series accumulation
    sqrtK = np.sqrt(K)

    for n in range(1, N + 1):
        beta = (2 * n - 1) * np.pi * sqrtK / (2 * L)
        odd = 2 * n - 1
        # common factors
        num1 = (4 / np.pi) / odd
        cterm = (beta**2 * nu + nu**2 + omega**2) * np.cos(omega * (tT - tau)) - (
            beta**2
        ) * omega * np.sin(omega * (tT - tau))
        denom = beta**4 + 2 * beta**2 * nu + nu**2 + omega**2

        total += (
            num1
            * ((nu * theta__M) / (beta**2 + nu) + theta__A * cterm / denom)
            * np.sin(beta / sqrtK * XX)
        )

    model = theta__A * np.cos(omega * (tT - tau)) + theta__M - total
    return model - TTT  # residuals (model - data)


def harmonic_model(V, tT, XX, TTT):
    """
    Residuals for boundary temperature model parameter estimation.
    V[0] = h  (via nu)
    V[1] = k  (via K)
    """
    theta__A = LBC_paramsA[0]
    theta__M = LBC_paramsA[1]
    tau = LBC_paramsA[2]
    h, k = V
    nu = (h * s) / (rho * A * c)  # convective term
    K = k / (rho * c)  # thermal diffusivity

    tT = np.asarray(tT, dtype=float)
    XX = np.asarray(XX, dtype=float)
    TTT = np.asarray(TTT, dtype=float)
    # Create an array of the size of tT to assign model values
    total = np.zeros_like(tT, dtype=float)

    # Series accumulation
    sqrtK = np.sqrt(K)
    for n in range(1, N + 1):
        beta = (2 * n - 1) * np.pi * sqrtK / (2 * L)
        odd = 2 * n - 1
        # common factors
        num1 = (4 / np.pi) / odd
        cterm = (beta**2 * nu + nu**2 + omega**2) * np.cos(omega * (tT - tau)) - (
            beta**2
        ) * omega * np.sin(omega * (tT - tau))
        denom = beta**4 + 2 * beta**2 * nu + nu**2 + omega**2

        total += (
            num1
            * ((nu * theta__M) / (beta**2 + nu) + theta__A * cterm / denom)
            * np.sin(beta / sqrtK * XX)
        )

    model = theta__A * np.cos(omega * (tT - tau)) + theta__M - total
    return model  # return model values


# Initial guess
x0 = [h0, h0]  # [h, k]

# Solve

harmonic_model_estimate = least_squares(harmonic_model_res, x0=x0, args=(tt, x, T))

# Best-fit parameters as array('d', ...)
harmonic_model_params = harmonic_model_estimate.x
harmonic_model_paramsA = arr.array("d", harmonic_model_params)  # Optimal parameters


for j in range(0, len(harmonic_model_params)):
    print(labels_col1[j + 7] + " = {}".format(harmonic_model_paramsA[j]))

h_value = harmonic_model_paramsA[0]
k_value = harmonic_model_paramsA[1]

best_hk[0] = float(h_value)
best_hk[1] = float(k_value)
sheet1.cell(row=7, column=1, value=labels_col1[6])
sheet1.cell(row=8, column=2, value=h_value)
sheet1.cell(row=9, column=2, value=k_value)
# ----- 95% Confidence Intervals from Jacobian -----
m = harmonic_model_estimate.fun.size
n = harmonic_model_estimate.x.size
dof = max(1, m - n)
print("Harmonic model degrees of freedom = ", dof)
RMS_harm_mod = np.sqrt(2 * harmonic_model_estimate.cost) / dof

print("Harmonic model RMS = ", RMS_harm_mod)
J = harmonic_model_estimate.jac
JTJ = J.T @ J
try:
    JTJ_inv = np.linalg.inv(JTJ)
except np.linalg.LinAlgError:
    JTJ_inv = np.linalg.pinv(JTJ)  # pseudo inverse if necessary

# cost = 0.5 * RSS  ->  RSS = 2*cost
sigma2 = (2.0 * harmonic_model_estimate.cost) / dof
cov_x = sigma2 * JTJ_inv
se = np.sqrt(np.diag(cov_x))

tcrit = scipy.stats.t.ppf(0.975, dof)
ci_lo = harmonic_model_params - tcrit * se
ci_hi = harmonic_model_params + tcrit * se

# Print CIs (h, k)
for j in range(n):
    print(f"{labels_col1[j+7]} 95% CI: [{ci_lo[j]}, {ci_hi[j]}]")

# Write the CI limits to the spreadsheet:
sheet1.cell(row=8, column=3, value=ci_lo[0])
sheet1.cell(row=8, column=4, value=ci_hi[0])
sheet1.cell(row=9, column=3, value=ci_lo[1])
sheet1.cell(row=9, column=4, value=ci_hi[1])

# save excel file
wb.save(str(excelsaveloc / f"{filesave}.xlsx"))

out_dir = Path("FIGS")
fname_base = "all_x_overlay"

# PNG for PPT slides:
png_path = plot_all_x_overlay(
    tt=tt,
    x=x,
    T=T,
    model_fn=harmonic_model,
    V=best_hk,
    out_dir=out_dir,
    fname_base=fname_base,
    plottype="PNG",
    dpi=300,
    show=True,
    model_kwargs=None,  # or omit if not needed
    XX_preferred=None,  # or e.g., [0.0, 5.0, 10.0, ...]
)

# PDF for publication:
pdf_path = plot_all_x_overlay(
    tt=tt,
    x=x,
    T=T,
    model_fn=harmonic_model,
    V=best_hk,
    out_dir=out_dir,
    fname_base=fname_base,
    plottype="PDF",
    dpi=300,
    show=False,
)
print("Wrote:", png_path, "and", pdf_path)
